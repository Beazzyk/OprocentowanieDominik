import openpyxl
location = r"C:\Users\domin\Downloads\Oprocentowanie.xlsx"
from_location = openpyxl.load_workbook("Oprocentowanie.xlsx")
months = from_location.active
print(months.cell(1, 1).value)

loan = 12000
inflation_1 = 1.592824484
inflation_2 = -0.453509101
inflation_3 = 2.324671717
inflation_4 = 1.261254407
inflation_5 = 1.782526286
inflation_6 = 2.329384541
inflation_7 = 1.502229842
inflation_8 = 1.782526286
inflation_9 = 2.328848994
inflation_10 = 0.616921348
inflation_11 = 2.352295886
inflation_12 = 0.337779545
inflation_13 = 1.577035247
inflation_14 = -0.292781443
inflation_15 = 2.48619659
inflation_16 = 0.267110318
inflation_17 = 1.417952672
inflation_18 = 1.054243267
inflation_19 = 1.480520104
inflation_20 = 1.577035247
inflation_21 = -0.07742069
inflation_22 = 1.165733399
inflation_23 = -0.404186718
inflation_24 = 1.499708521

interest_rate = float(input("Podaj oprocentowanie kredytu: "))
original_amount = float(input("Podaj początkowa kwotę kredytu: "))
fixed_instalment = float(input("Podaj stała wartość raty: "))
print(f"Oprocentowanie to: {interest_rate} %,"f"Początkowa wartość kredytu to: {original_amount} zl," f"Stała wartość raty to: {fixed_instalment} zl")

for row in range(2, months.max_row + 1):
    inflation_cell = months.cell(row, 2)
    if inflation_cell.value is None:
        continue
    inflation = float(inflation_cell.value)
    remaining_amount = (1 + ((inflation + interest_rate) / 1200)) * loan - fixed_instalment
    difference = loan - remaining_amount
    print(f"Twoja pozostała kwota kredytu to {remaining_amount:.2f} zł, to {difference:.2f} zł mniej niż w poprzednim miesiącu.")
    previous_amount = remaining_amount

first_instalment = (1+((inflation_1 + interest_rate) / 1200)) * loan - fixed_instalment
second_instalment = (1+((inflation_2 + interest_rate) / 1200)) * loan - fixed_instalment
third_instalment = (1+((inflation_3 + interest_rate) / 1200)) * loan - fixed_instalment
fourth_instalment = (1+((inflation_4 + interest_rate) / 1200)) * loan - fixed_instalment
fifth_instalment = (1+((inflation_5 + interest_rate) / 1200)) * loan - fixed_instalment
sixth_instalment = (1+((inflation_6 + interest_rate) / 1200)) * loan - fixed_instalment
seventh_instalment = (1+((inflation_7 + interest_rate) / 1200)) * loan - fixed_instalment
eighth_instalment = (1+((inflation_8 + interest_rate) / 1200)) * loan - fixed_instalment
ninth_instalment = (1+((inflation_9 + interest_rate) / 1200)) * loan - fixed_instalment
tenth_instalment = (1+((inflation_10 + interest_rate) / 1200)) * loan - fixed_instalment
eleventh_instalment = (1+((inflation_11 + interest_rate) / 1200)) * loan - fixed_instalment
twelfth_instalment = (1+((inflation_12 + interest_rate) / 1200)) * loan - fixed_instalment
thirteenth_instalment = (1+((inflation_13 + interest_rate) / 1200)) * loan - fixed_instalment
fourteenth_instalment = (1+((inflation_14 + interest_rate) / 1200)) * loan - fixed_instalment
fifteenth_instalment = (1+((inflation_15 + interest_rate) / 1200)) * loan - fixed_instalment
sixteenth_instalment = (1+((inflation_16 + interest_rate) / 1200)) * loan - fixed_instalment
seventeenth_instalment = (1+((inflation_17 + interest_rate) / 1200)) * loan - fixed_instalment
eighteen_instalment = (1+((inflation_18 + interest_rate) / 1200)) * loan - fixed_instalment
nineteen_instalment = (1+((inflation_19 + interest_rate) / 1200)) * loan - fixed_instalment
twenty_instalment = (1+((inflation_20 + interest_rate) / 1200)) * loan - fixed_instalment
twentyfirst_instalment = (1+((inflation_21 + interest_rate) / 1200)) * loan - fixed_instalment
twentytwo_instalment = (1+((inflation_22 + interest_rate) / 1200)) * loan - fixed_instalment
twentythree_instalment = (1+((inflation_23 + interest_rate) / 1200)) * loan - fixed_instalment
twentyfour_instalment = (1+((inflation_24 + interest_rate) / 1200)) * loan - fixed_instalment

print(f"Twoja pozostała kwota kredytu to {first_instalment:.2f} zł, to {loan - first_instalment:.2f} zł mniej niż w poprzednim miesiącu.")


# (1+((inflation_1+interest_rate)/1200)) * loan - fixed_instalment
#